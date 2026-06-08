#!/usr/bin/env python3
"""
generate_prd_review.py
生成 PRD 质量评审 Excel，包含默认处理建议供用户确认。

使用方式:
    python generate_prd_review.py <review_json> <output_xlsx>

输入 JSON schema:
{
    "project": "项目名",
    "prd_source": "PRD文件名",
    "findings": [
        {
            "id": "PR-001",
            "severity": "blocker",          // blocker / major / minor
            "dimension": "缺失",             // 缺失/矛盾/模糊/不可测/奇脉/依赖
            "location": "AC-2.1.2",         // PRD中的位置
            "quote": "单文件大小限制 50MB。超出时前端禁用上传并显示提示",
            "problem": "问题的具体描述",
            "impact": "该问题对测试用例的影响",
            "default_action": "补全",        // 补全/假定/搁置/追问PM
            "default_detail": "按通用实践同时测前端（JS拦截）+后端（413响应）两个场景",
            "default_new_cases": [           // 若用户确认，会按此加入的新用例大纲
                "TC-NEW: 文件大小恰好50MB通过",
                "TC-NEW: 文件大小50MB+1字节前端拦截",
                "TC-NEW: 绕过前端的60MB请求被后端返回413"
            ]
        }
    ]
}
"""

import json
import sys
from pathlib import Path
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="305496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

SEVERITY_FILLS = {
    "blocker": PatternFill("solid", start_color="FFC7CE"),
    "major": PatternFill("solid", start_color="FFEB9C"),
    "minor": PatternFill("solid", start_color="C6EFCE"),
}

SEVERITY_DISPLAY = {
    "blocker": "🔴 阻塞",
    "major": "🟡 重要",
    "minor": "🟢 建议",
}

DIMENSION_COLORS = {
    "缺失": "BDD7EE",
    "矛盾": "F4B084",
    "模糊": "FFE699",
    "不可测": "C9C9C9",
    "奇脉": "B4C7E7",
    "依赖": "D5A6BD",
}

ACTION_DISPLAY_MAP = {
    "补全": "补全（用例中按建议生成）",
    "假定": "假定（采用建议默认值）",
    "搁置": "搁置（该场景暂不测）",
    "追问PM": "追问 PM（必须先明确）",
    "追问": "追问 PM（必须先明确）",
}

COLUMNS = [
    ("用户决定", "user_decision", 14),   # 用户勾选列
    ("问题 ID", "id", 10),
    ("严重等级", "severity_display", 10),
    ("维度", "dimension", 10),
    ("PRD 位置", "location", 14),
    ("原文引用", "quote", 28),
    ("问题说明", "problem", 36),
    ("对测试的影响", "impact", 30),
    ("默认处理建议", "default_action_display", 18),
    ("建议详情", "default_detail", 36),
    ("若确认将新增的用例", "default_new_cases", 40),
    ("用户补充/覆盖说明", "user_notes", 24),
]


def format_new_cases(cases_list):
    if not cases_list:
        return "（无新增用例）"
    return "\n".join(f"• {c}" for c in cases_list)


def write_findings_sheet(wb, data):
    ws = wb.active
    ws.title = "PRD评审"

    # 说明行
    ws["A1"] = (
        "✍️ 使用说明：在【用户决定】列填'接受'/'修改'/'拒绝'。"
        "若选'修改'，请在【用户补充/覆盖说明】列写明新的处理方式。全部确认后另存为 xlsx，作为阶段A生成用例的输入。"
    )
    ws["A1"].font = Font(name="Arial", size=10, italic=True, color="C00000")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    # 表头（第 2 行）
    for col_idx, (header, _, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    findings = data.get("findings", [])

    for row_offset, f in enumerate(findings):
        row_idx = row_offset + 3
        severity = f.get("severity", "").lower()
        dimension = f.get("dimension", "")
        action = f.get("default_action", "").strip()
        action_display = ACTION_DISPLAY_MAP.get(action, action)

        values = {
            "user_decision": "",
            "id": f.get("id", ""),
            "severity_display": SEVERITY_DISPLAY.get(severity, severity),
            "dimension": dimension,
            "location": f.get("location", ""),
            "quote": f.get("quote", ""),
            "problem": f.get("problem", ""),
            "impact": f.get("impact", ""),
            "default_action_display": action_display,
            "default_detail": f.get("default_detail", ""),
            "default_new_cases": format_new_cases(f.get("default_new_cases", [])),
            "user_notes": "",
        }

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=values[key])
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER

            if key == "severity_display":
                fill = SEVERITY_FILLS.get(severity)
                if fill:
                    cell.fill = fill
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    cell.font = Font(name="Arial", size=10, bold=True)

            if key == "dimension":
                color = DIMENSION_COLORS.get(dimension)
                if color:
                    cell.fill = PatternFill("solid", start_color=color)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            if key in ("id", "location"):
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            if key == "user_decision":
                cell.fill = PatternFill("solid", start_color="FFF2CC")
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            if key == "user_notes":
                cell.fill = PatternFill("solid", start_color="FFF9E6")

    # 列宽
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 行高
    for row_idx in range(3, len(findings) + 3):
        ws.row_dimensions[row_idx].height = 120

    # 冻结到 C 列，便于左右滚动仍见 ID 和严重等级
    ws.freeze_panes = "D3"

    # 筛选器
    last_col_letter = get_column_letter(len(COLUMNS))
    last_row = len(findings) + 2
    ws.auto_filter.ref = f"A2:{last_col_letter}{last_row}"

    # 用户决定列的下拉
    if findings:
        dv = DataValidation(
            type="list",
            formula1='"接受,修改,拒绝"',
            allow_blank=True,
        )
        dv.prompt = "接受=按建议执行\n修改=在补充列说明\n拒绝=不处理该问题"
        dv.promptTitle = "如何处理"
        dv.add(f"A3:A{len(findings) + 2}")
        ws.add_data_validation(dv)


def write_summary_sheet(wb, data):
    ws = wb.create_sheet("评审摘要", 0)
    findings = data.get("findings", [])

    ws["A1"] = f"PRD 质量评审 - {data.get('project', '未命名')}"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="305496")
    ws.merge_cells("A1:D1")

    ws["A2"] = f"PRD 来源：{data.get('prd_source', '未指定')}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="595959")
    ws.merge_cells("A2:D2")

    total = len(findings)
    by_severity = Counter(f.get("severity", "").lower() for f in findings)
    by_dimension = Counter(f.get("dimension", "") for f in findings)
    by_action = Counter(f.get("default_action", "") for f in findings)

    row = 4
    ws.cell(row=row, column=1, value="总览").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    ws.cell(row=row, column=1, value="问题总数").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total)
    row += 2

    def write_block(title, counter, color_map=None):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        ws.cell(row=row, column=1, value="分类").font = Font(bold=True)
        ws.cell(row=row, column=2, value="数量").font = Font(bold=True)
        ws.cell(row=row, column=3, value="占比").font = Font(bold=True)
        row += 1
        for k, v in counter.most_common():
            if not k:
                continue
            display = SEVERITY_DISPLAY.get(k, k) if title == "按严重等级" else k
            ws.cell(row=row, column=1, value=display)
            ws.cell(row=row, column=2, value=v).alignment = Alignment(horizontal="center")
            pct = f"{(v / total * 100):.1f}%" if total else "0%"
            ws.cell(row=row, column=3, value=pct).alignment = Alignment(horizontal="center")
            if color_map and k in color_map:
                ws.cell(row=row, column=1).fill = color_map[k]
            row += 1
        row += 1

    write_block("按严重等级", by_severity, SEVERITY_FILLS)
    write_block("按维度", by_dimension)
    write_block("按默认处理建议", by_action)

    # 使用提示
    row += 1
    ws.cell(row=row, column=1, value="下一步操作").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    steps = [
        "1. 打开【PRD评审】Sheet，逐条审视问题",
        "2. 在【用户决定】列选择：接受 / 修改 / 拒绝",
        "3. 选'修改'时，在【用户补充/覆盖说明】列写明新方案",
        "4. 全部确认后另存为 xlsx，作为阶段 A 生成用例的输入",
        "5. Skill 会自动根据你的决定，把'接受'和'修改'对应的场景加入用例表",
    ]
    for step in steps:
        ws.cell(row=row, column=1, value=step).font = BODY_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20


def main():
    if len(sys.argv) != 3:
        print("用法: python generate_prd_review.py <review.json> <output.xlsx>")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    write_findings_sheet(wb, data)
    write_summary_sheet(wb, data)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    n = len(data.get("findings", []))
    print(f"✅ 已生成 {n} 条 PRD 评审问题 → {output_path}")
    print("   下一步：用户在 Excel 的【用户决定】列填写处理方式，另存后可进入阶段 A")


if __name__ == "__main__":
    main()

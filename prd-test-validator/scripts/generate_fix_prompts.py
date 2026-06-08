#!/usr/bin/env python3
"""
generate_fix_prompts.py
读取用户已勾选的问题清单 Excel（修复列填了 √ 的行），按文件分组，
生成三种格式的修复指令：
    - fix_prompt_generic.md       通用 Markdown，可粘贴到任何 AI 工具
    - fix_prompt_cursor/          Cursor 专用目录（.cursorrules + 分条任务）
    - fix_prompt_claude_code.md   Claude Code 适用的分批任务清单

使用方式:
    python generate_fix_prompts.py <checked_xlsx> <output_dir>
    python generate_fix_prompts.py <checked_xlsx> <output_dir> --project "项目名"
"""

import argparse
import sys
from pathlib import Path
from collections import defaultdict, Counter

from openpyxl import load_workbook


TRUTHY_FLAGS = {"√", "x", "X", "✓", "✔", "yes", "Yes", "YES", "1", "true", "True", "是"}

SEVERITY_ORDER = {"critical": 0, "major": 1, "minor": 2}
SEVERITY_DISPLAY = {
    "critical": "🔴 严重",
    "major": "🟡 一般",
    "minor": "🟢 建议",
}


def read_checked_issues(xlsx_path):
    """读取勾选的问题，返回字典列表。"""
    wb = load_workbook(xlsx_path, data_only=True)
    # 优先找"问题清单" sheet；否则找包含"问题"或"issue"的 sheet
    ws = None
    for name in wb.sheetnames:
        if name == "问题清单":
            ws = wb[name]
            break
    if ws is None:
        for name in wb.sheetnames:
            if "问题" in name or "issue" in name.lower():
                ws = wb[name]
                break
    if ws is None:
        raise ValueError("找不到'问题清单' sheet，请确认输入 Excel 是 generate_issues_checklist.py 的输出")

    # 表头在第 2 行
    headers = [cell.value for cell in ws[2]]
    header_map = {h: idx for idx, h in enumerate(headers)}

    required_cols = ["修复", "问题 ID", "严重等级", "维度", "文件", "行号",
                     "问题标题", "代码片段", "问题说明", "修复建议"]
    missing = [c for c in required_cols if c not in header_map]
    if missing:
        raise ValueError(f"Excel 表头缺少列: {missing}")

    issues = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or all(c is None for c in row):
            continue

        flag_val = row[header_map["修复"]]
        flag_str = str(flag_val).strip() if flag_val is not None else ""
        if flag_str not in TRUTHY_FLAGS:
            continue

        sev_display = str(row[header_map["严重等级"]] or "").strip()
        # 从展示名反解回标准值
        sev = "minor"
        for key, disp in SEVERITY_DISPLAY.items():
            if disp in sev_display or key in sev_display.lower():
                sev = key
                break

        issue = {
            "id": str(row[header_map["问题 ID"]] or "").strip(),
            "severity": sev,
            "dimension": str(row[header_map["维度"]] or "").strip(),
            "file": str(row[header_map["文件"]] or "").strip(),
            "line": str(row[header_map["行号"]] or "").strip(),
            "title": str(row[header_map["问题标题"]] or "").strip(),
            "code_snippet": str(row[header_map["代码片段"]] or "").strip(),
            "description": str(row[header_map["问题说明"]] or "").strip(),
            "fix_suggestion": str(row[header_map["修复建议"]] or "").strip(),
            "related_ac": str(row[header_map.get("关联 AC", -1)] or "").strip() if "关联 AC" in header_map else "",
            "related_tc": str(row[header_map.get("关联用例", -1)] or "").strip() if "关联用例" in header_map else "",
        }
        issues.append(issue)

    return issues


def group_by_file(issues):
    """按文件分组，每组内按严重等级排序。"""
    groups = defaultdict(list)
    for i in issues:
        groups[i["file"]].append(i)
    for file, items in groups.items():
        items.sort(key=lambda x: (SEVERITY_ORDER.get(x["severity"], 99), x["id"]))
    return groups


# ============ 格式 1: 通用 Markdown ============

GENERIC_HEADER = """# 代码修复任务（自动生成）

项目：{project}
共 **{n_issues}** 个问题分布在 **{n_files}** 个文件，按文件组织成 {n_files} 个任务。
粘贴本文件到你的 AI 编码工具（Cursor/Claude Code/Windsurf/Copilot Chat 等）即可。

---

## 给 AI 的通用指令（请先阅读）

- 严格按照下方的问题列表修复，**不要擅自改动无关代码**。
- 每个问题修改后，在原代码位置附近保留一行注释：`# fixed: <问题ID>`（或对应语言的注释语法）。
- 如果某个问题的修复需要引入新依赖（库、包），在任务结束时列出所有新增依赖。
- 如果遇到拿不准的修复方向，**不要猜，先列出疑问再停下**。
- 所有修复应尽量保持原有代码风格和命名习惯。
- 修复完成后，自测：代码能被 import / 语法无误 / 没有引入明显的新 bug。

---
"""

GENERIC_FILE_SECTION = """
## 任务 {task_num} / {total_tasks}：修复 `{file}`

**本文件共 {n} 个问题** — 🔴 严重 {critical}，🟡 一般 {major}，🟢 建议 {minor}

{issue_blocks}

---
"""

GENERIC_ISSUE_BLOCK = """### [{issue_id}] {severity_display} {title}

- **维度**：{dimension}
- **位置**：`{file}:{line}`
- **问题**：{description}

**当前代码：**
```
{code_snippet}
```

**修复建议：**
{fix_suggestion}
"""


def render_generic(groups, project):
    total_issues = sum(len(v) for v in groups.values())
    n_files = len(groups)
    parts = [GENERIC_HEADER.format(project=project, n_issues=total_issues, n_files=n_files)]

    for idx, (file, issues) in enumerate(sorted(groups.items()), start=1):
        sev_count = Counter(i["severity"] for i in issues)
        blocks = []
        for issue in issues:
            blocks.append(GENERIC_ISSUE_BLOCK.format(
                issue_id=issue["id"],
                severity_display=SEVERITY_DISPLAY.get(issue["severity"], ""),
                title=issue["title"],
                dimension=issue["dimension"],
                file=issue["file"],
                line=issue["line"],
                description=issue["description"],
                code_snippet=issue["code_snippet"] or "(无具体片段)",
                fix_suggestion=issue["fix_suggestion"],
            ))
        parts.append(GENERIC_FILE_SECTION.format(
            task_num=idx,
            total_tasks=n_files,
            file=file,
            n=len(issues),
            critical=sev_count.get("critical", 0),
            major=sev_count.get("major", 0),
            minor=sev_count.get("minor", 0),
            issue_blocks="\n\n".join(blocks),
        ))

    parts.append("\n## 修复完成后\n\n- 逐文件 diff 自检，确保修改合理\n- 运行项目已有的测试（如有）\n- 把修复的问题 ID 列回给调用方以便对照验收\n")
    return "\n".join(parts)


# ============ 格式 2: Cursor 专用 ============

CURSORRULES_TEMPLATE = """# Cursor Rules - 代码修复任务

你是一位严谨的代码审查修复工程师。以下规则在本次修复会话中必须遵守：

## 修复原则
1. 严格按照 `.cursor-tasks/` 目录下的任务文件逐个修复，不要擅自改动无关代码。
2. 修复优先级：🔴 严重 > 🟡 一般 > 🟢 建议。如果时间/上下文有限，优先修🔴。
3. 每处修改在原代码附近留下 `# fixed: <问题ID>`（或对应语言的注释语法）。
4. 同文件的多个问题合并在一次编辑里完成，避免来回切换。
5. 涉及的新依赖（库、包、配置）在每个任务末尾列清单。

## 禁止
- 不要重构无关代码
- 不要为了"看起来更好"改变量名、函数签名
- 不要引入新的抽象层（class/wrapper）除非修复必需

## 遇到问题时
- 不确定修复方向时，先在聊天里列出 2-3 个方案让用户选
- 发现修复会破坏已有测试时，先报告再改

## 项目上下文
项目：{project}
共 {n_issues} 个问题待修复，分布在 {n_files} 个文件。
"""

CURSOR_TASK_TEMPLATE = """# 任务 {task_num}: 修复 {file}

- 问题数：{n}（🔴 {critical} / 🟡 {major} / 🟢 {minor}）
- 严格按顺序处理下方每个问题

{issue_blocks}

## 完成标准
- [ ] 每个问题已处理（已修复 或 已用注释说明为什么不修）
- [ ] 代码能正常 import / 通过语法检查
- [ ] 在文件末尾追加注释列出本次修复了哪些问题 ID
"""

CURSOR_TASK_ISSUE = """## [{issue_id}] {severity_display} {title}

**位置**: `{file}:{line}` · **维度**: {dimension}

**问题描述**：{description}

**当前代码片段**：
```
{code_snippet}
```

**修复方向**：
{fix_suggestion}
"""


def render_cursor(groups, project, output_dir):
    """Cursor 专用：生成 .cursorrules + .cursor-tasks/ 下的分文件任务"""
    output_dir = Path(output_dir)
    cursor_dir = output_dir / "fix_prompt_cursor"
    cursor_dir.mkdir(parents=True, exist_ok=True)

    total_issues = sum(len(v) for v in groups.values())
    n_files = len(groups)

    # .cursorrules
    rules_content = CURSORRULES_TEMPLATE.format(
        project=project, n_issues=total_issues, n_files=n_files
    )
    (cursor_dir / ".cursorrules").write_text(rules_content, encoding="utf-8")

    # tasks 目录
    tasks_dir = cursor_dir / ".cursor-tasks"
    tasks_dir.mkdir(exist_ok=True)

    # 索引文件
    index_lines = ["# Cursor 修复任务索引\n",
                   f"总计 {total_issues} 个问题，{n_files} 个任务。按顺序执行：\n"]

    for idx, (file, issues) in enumerate(sorted(groups.items()), start=1):
        sev_count = Counter(i["severity"] for i in issues)
        safe_name = file.replace("/", "_").replace("\\", "_").replace(" ", "_")
        task_filename = f"task_{idx:02d}_{safe_name}.md"

        blocks = []
        for issue in issues:
            blocks.append(CURSOR_TASK_ISSUE.format(
                issue_id=issue["id"],
                severity_display=SEVERITY_DISPLAY.get(issue["severity"], ""),
                title=issue["title"],
                file=issue["file"],
                line=issue["line"],
                dimension=issue["dimension"],
                description=issue["description"],
                code_snippet=issue["code_snippet"] or "(无具体片段)",
                fix_suggestion=issue["fix_suggestion"],
            ))

        content = CURSOR_TASK_TEMPLATE.format(
            task_num=idx, file=file, n=len(issues),
            critical=sev_count.get("critical", 0),
            major=sev_count.get("major", 0),
            minor=sev_count.get("minor", 0),
            issue_blocks="\n\n".join(blocks),
        )
        (tasks_dir / task_filename).write_text(content, encoding="utf-8")

        index_lines.append(f"{idx}. [{task_filename}](./.cursor-tasks/{task_filename}) - `{file}`（{len(issues)} 个问题）")

    index_lines.append("\n## 使用方法\n")
    index_lines.append("1. 把本目录（含 `.cursorrules`）复制到项目根目录")
    index_lines.append("2. 在 Cursor 中打开项目，Cursor 会自动读取 `.cursorrules`")
    index_lines.append("3. 用 Cmd+L 打开对话，发送：`请开始执行 .cursor-tasks/task_01_xxx.md`")
    index_lines.append("4. 修完一个任务后发送下一个")
    index_lines.append("\n也可以一次性：`请依次执行 .cursor-tasks/ 下的所有任务文件`（谨慎，上下文较长时 Cursor 可能漏掉问题）")

    (cursor_dir / "README.md").write_text("\n".join(index_lines), encoding="utf-8")
    return cursor_dir


# ============ 格式 3: Claude Code ============

CLAUDE_CODE_HEADER = """# Claude Code 修复任务清单

项目：{project}
总计 {n_issues} 个问题，{n_files} 个文件。

## 使用方法

把本文件作为 prompt 发给 Claude Code（`claude` CLI 或 Claude 桌面端的 Code mode）。Claude Code 会按顺序处理每个文件，并给出 diff 供你审阅。

推荐启动命令（在项目根目录）：
```bash
claude "请按照 fix_prompt_claude_code.md 的任务清单，使用 Edit 工具逐文件修复。每修完一个文件先给我 diff 再进下一个文件。"
```

## 修复原则（请 Claude 严格遵守）

- 使用 Edit/MultiEdit 工具精确修改，保留所有无关代码原样
- 每处修改在附近加一行注释 `# fixed: <问题ID>`
- 同一文件的多个问题，尽量在一次 MultiEdit 调用里完成，减少往返
- 每完成一个文件，用 git diff 检查，然后停下来等用户确认
- 如果需要新依赖，单独列在回复末尾

---
"""

CLAUDE_CODE_TASK = """## 任务 {task_num}/{total}: `{file}`

共 {n} 个问题（🔴 {critical} / 🟡 {major} / 🟢 {minor}）

{issue_blocks}

**完成标准**：所有上述问题在本文件内都已处理，且文件能通过语法检查。

---
"""

CLAUDE_CODE_ISSUE = """### {issue_id} {severity_display} {title}

- 位置：`{file}:{line}`
- 维度：{dimension}
- 问题：{description}

```
{code_snippet}
```

修复方向：{fix_suggestion}
"""


def render_claude_code(groups, project):
    total_issues = sum(len(v) for v in groups.values())
    n_files = len(groups)
    parts = [CLAUDE_CODE_HEADER.format(
        project=project, n_issues=total_issues, n_files=n_files
    )]

    for idx, (file, issues) in enumerate(sorted(groups.items()), start=1):
        sev_count = Counter(i["severity"] for i in issues)
        blocks = []
        for issue in issues:
            blocks.append(CLAUDE_CODE_ISSUE.format(
                issue_id=issue["id"],
                severity_display=SEVERITY_DISPLAY.get(issue["severity"], ""),
                title=issue["title"],
                file=issue["file"],
                line=issue["line"],
                dimension=issue["dimension"],
                description=issue["description"],
                code_snippet=issue["code_snippet"] or "(无具体片段)",
                fix_suggestion=issue["fix_suggestion"],
            ))
        parts.append(CLAUDE_CODE_TASK.format(
            task_num=idx, total=n_files, file=file, n=len(issues),
            critical=sev_count.get("critical", 0),
            major=sev_count.get("major", 0),
            minor=sev_count.get("minor", 0),
            issue_blocks="\n".join(blocks),
        ))

    parts.append("\n## 全部完成后\n")
    parts.append("1. 运行 `git diff --stat` 列出所有改动")
    parts.append("2. 列出本次修复处理的所有问题 ID")
    parts.append("3. 列出新增依赖（如有）和需要在环境变量或配置文件中调整的项")
    return "\n".join(parts)


# ============ 主流程 ============

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", help="用户已勾选的问题清单 Excel 路径")
    parser.add_argument("output_dir", help="输出目录")
    parser.add_argument("--project", default="未命名项目", help="项目名")
    args = parser.parse_args()

    issues = read_checked_issues(args.xlsx)
    if not issues:
        print("⚠️ Excel 中没有任何行被勾选（修复列为空或值不在 {√, x, yes, 1} 中）")
        sys.exit(1)

    groups = group_by_file(issues)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 格式 1: 通用 Markdown
    generic_path = output_dir / "fix_prompt_generic.md"
    generic_path.write_text(render_generic(groups, args.project), encoding="utf-8")
    print(f"✅ 通用 Markdown: {generic_path}")

    # 格式 2: Cursor
    cursor_dir = render_cursor(groups, args.project, output_dir)
    print(f"✅ Cursor 专用目录: {cursor_dir}/")

    # 格式 3: Claude Code
    cc_path = output_dir / "fix_prompt_claude_code.md"
    cc_path.write_text(render_claude_code(groups, args.project), encoding="utf-8")
    print(f"✅ Claude Code 任务清单: {cc_path}")

    # 汇总
    sev_count = Counter(i["severity"] for i in issues)
    print()
    print(f"📊 共收集 {len(issues)} 个勾选问题，分布在 {len(groups)} 个文件：")
    print(f"   🔴 严重 {sev_count.get('critical', 0)}  🟡 一般 {sev_count.get('major', 0)}  🟢 建议 {sev_count.get('minor', 0)}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
generate_test_cases.py
从 JSON 输入生成格式化的测试用例 Excel 文件。

使用方式:
    python generate_test_cases.py <cases_json_path> <output_xlsx_path>

输入 JSON schema:
{
    "project": "项目名称",
    "prd_source": "PRD来源（文件名或描述）",
    "test_cases": [
        {
            "id": "TC-LOGIN-001",
            "module": "登录模块",
            "title": "使用有效账号密码登录成功",
            "priority": "P0",                      // P0 / P1 / P2
            "case_type": "正向",                    // 正向 / 负向 / 异常
            "design_method": "等价类+场景法",        // 等价类/边界值/因果图/场景法/错误推测
            "precondition": "1. 已注册账号 user1/pass123\n2. 浏览器已清除cookie",
            "steps": "1. 打开登录页\n2. 输入账号 user1\n3. 输入密码 pass123\n4. 点击登录",
            "test_data": "账号: user1, 密码: pass123",
            "expected": "1. 跳转至首页\n2. 右上角显示用户名\n3. cookie 中写入 session_id",
            "test_type": "功能",                    // 功能 / 性能 / 安全 / 兼容
            "related_ac": "AC-1.1"
        }
    ],
    "compatibility_matrix": {
        "browsers": ["Chrome 最新版", "Firefox 最新版", ...],
        "os": ["Windows 11", "macOS 最新版", ...],
        "devices": ["iPhone 15", "Samsung S24", ...],
        "resolutions": ["1920x1080", "1366x768"],
        "notes": "兼容性要求说明"
    }
}
"""

import json
import sys
from pathlib import Path
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============ 样式常量 ============
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="305496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

PRIORITY_FILLS = {
    "P0": PatternFill("solid", start_color="FFC7CE"),  # 淡红
    "P1": PatternFill("solid", start_color="FFEB9C"),  # 淡黄
    "P2": PatternFill("solid", start_color="C6EFCE"),  # 淡绿
}

CASE_TYPE_FONT_COLORS = {
    "正向": "006100",
    "负向": "9C0006",
    "异常": "9C5700",
}


# ============ 列定义 ============
COLUMNS = [
    ("用例 ID", "id", 16),
    ("所属模块", "module", 14),
    ("用例标题", "title", 36),
    ("优先级", "priority", 8),
    ("用例类型", "case_type", 10),
    ("测试设计方法", "design_method", 16),
    ("前置条件", "precondition", 28),
    ("测试步骤", "steps", 40),
    ("测试数据", "test_data", 22),
    ("预期结果", "expected", 36),
    ("测试类型", "test_type", 10),
    ("关联验收标准", "related_ac", 14),
]


def write_test_cases_sheet(wb, data):
    """写入测试用例主 sheet"""
    ws = wb.active
    ws.title = "测试用例"

    # 写表头
    headers = [col[0] for col in COLUMNS]
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 写数据
    cases = data.get("test_cases", [])
    for case in cases:
        row = [case.get(col[1], "") for col in COLUMNS]
        ws.append(row)

    # 样式：逐行设置
    for row_idx in range(2, len(cases) + 2):
        for col_idx, (header, key, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER

            # 优先级着色
            if key == "priority":
                p = str(cell.value).strip().upper()
                if p in PRIORITY_FILLS:
                    cell.fill = PRIORITY_FILLS[p]
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    cell.font = Font(name="Arial", size=10, bold=True)

            # 用例类型着色
            if key == "case_type":
                ct = str(cell.value).strip()
                color = CASE_TYPE_FONT_COLORS.get(ct)
                if color:
                    cell.font = Font(name="Arial", size=10, bold=True, color=color)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            # ID、测试类型、关联AC居中
            if key in ("id", "test_type", "related_ac"):
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

    # 列宽
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 行高自适应：让长文本行更高
    for row_idx in range(2, len(cases) + 2):
        ws.row_dimensions[row_idx].height = 60

    # 冻结首行
    ws.freeze_panes = "A2"

    # 筛选器（给整个数据区加 auto filter）
    last_col_letter = get_column_letter(len(COLUMNS))
    last_row = len(cases) + 1
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"


def write_compatibility_sheet(wb, data):
    """写入兼容性矩阵 sheet"""
    ws = wb.create_sheet("兼容性矩阵")
    compat = data.get("compatibility_matrix", {})

    # 标题行
    ws["A1"] = "兼容性测试矩阵"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="305496")
    ws.merge_cells("A1:D1")

    if compat.get("notes"):
        ws["A2"] = f"说明：{compat['notes']}"
        ws["A2"].font = Font(name="Arial", size=10, italic=True, color="595959")
        ws.merge_cells("A2:D2")

    row = 4
    sections = [
        ("浏览器", compat.get("browsers", [])),
        ("操作系统", compat.get("os", [])),
        ("设备", compat.get("devices", [])),
        ("分辨率", compat.get("resolutions", [])),
    ]

    for section_title, items in sections:
        if not items:
            continue
        ws.cell(row=row, column=1, value=section_title).font = Font(
            name="Arial", size=11, bold=True, color="FFFFFF"
        )
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        for idx, item in enumerate(items, start=1):
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(
                horizontal="center"
            )
            ws.cell(row=row, column=2, value=item).font = BODY_FONT
            ws.cell(row=row, column=3, value="待测试").alignment = Alignment(
                horizontal="center"
            )
            ws.cell(row=row, column=4, value="").alignment = BODY_ALIGN

            for c in range(1, 5):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

        row += 1  # 空行分组

    # 列标题
    ws.cell(row=3, column=1, value="序号").font = HEADER_FONT
    ws.cell(row=3, column=2, value="项").font = HEADER_FONT
    ws.cell(row=3, column=3, value="状态").font = HEADER_FONT
    ws.cell(row=3, column=4, value="备注").font = HEADER_FONT
    for c in range(1, 5):
        cell = ws.cell(row=3, column=c)
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40
    ws.freeze_panes = "A4"


def write_summary_sheet(wb, data):
    """写入用例索引/统计 sheet"""
    ws = wb.create_sheet("用例索引", 0)  # 放到最前
    cases = data.get("test_cases", [])

    ws["A1"] = f"测试用例统计 - {data.get('project', '未命名项目')}"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="305496")
    ws.merge_cells("A1:D1")

    ws["A2"] = f"PRD 来源：{data.get('prd_source', '未指定')}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="595959")
    ws.merge_cells("A2:D2")

    # 统计
    total = len(cases)
    by_priority = Counter(c.get("priority", "").upper() for c in cases)
    by_type = Counter(c.get("case_type", "") for c in cases)
    by_test_type = Counter(c.get("test_type", "") for c in cases)
    by_module = Counter(c.get("module", "") for c in cases)

    row = 4

    # 总览
    ws.cell(row=row, column=1, value="总览").font = Font(
        name="Arial", size=12, bold=True, color="FFFFFF"
    )
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    ws.cell(row=row, column=1, value="用例总数").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total)
    row += 2

    def write_stat_block(title, counter):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = Font(
            name="Arial", size=12, bold=True, color="FFFFFF"
        )
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        ws.cell(row=row, column=1, value="分类").font = Font(bold=True)
        ws.cell(row=row, column=2, value="数量").font = Font(bold=True)
        ws.cell(row=row, column=3, value="占比").font = Font(bold=True)
        row += 1
        for k, v in counter.most_common():
            if not k:
                continue
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v).alignment = Alignment(
                horizontal="center"
            )
            pct = f"{(v / total * 100):.1f}%" if total else "0%"
            ws.cell(row=row, column=3, value=pct).alignment = Alignment(
                horizontal="center"
            )
            if title == "按优先级" and k in PRIORITY_FILLS:
                ws.cell(row=row, column=1).fill = PRIORITY_FILLS[k]
            row += 1
        row += 1

    write_stat_block("按优先级", by_priority)
    write_stat_block("按用例类型", by_type)
    write_stat_block("按测试类型", by_test_type)
    write_stat_block("按模块", by_module)

    # 列宽
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20


def validate_cases(data):
    """生成前做健康检查，把问题打印出来（不阻断，但提醒）"""
    cases = data.get("test_cases", [])
    warnings = []

    if not cases:
        warnings.append("❌ 没有测试用例")
        return warnings

    # 检查 ID 重复
    ids = [c.get("id") for c in cases]
    dup = [i for i, cnt in Counter(ids).items() if cnt > 1]
    if dup:
        warnings.append(f"⚠️ 发现重复的用例 ID: {dup}")

    # 检查必填字段
    required = ["id", "module", "title", "priority", "steps", "expected"]
    for idx, c in enumerate(cases):
        missing = [f for f in required if not c.get(f)]
        if missing:
            warnings.append(f"⚠️ 用例 #{idx+1} ({c.get('id','无ID')}) 缺字段: {missing}")

    # 检查优先级值
    valid_priorities = {"P0", "P1", "P2"}
    for c in cases:
        if c.get("priority", "").upper() not in valid_priorities:
            warnings.append(
                f"⚠️ 用例 {c.get('id')} 的优先级 '{c.get('priority')}' 无效，应为 P0/P1/P2"
            )

    # 检查 P0 比例
    p0_count = sum(1 for c in cases if c.get("priority", "").upper() == "P0")
    p0_ratio = p0_count / len(cases)
    if p0_ratio > 0.5:
        warnings.append(
            f"⚠️ P0 占比 {p0_ratio:.1%} 偏高（建议 20-40%），考虑重新评估优先级"
        )

    # 检查每个模块是否有正+负用例
    module_types = {}
    for c in cases:
        m = c.get("module", "")
        t = c.get("case_type", "")
        module_types.setdefault(m, set()).add(t)
    for m, types in module_types.items():
        if "正向" not in types:
            warnings.append(f"⚠️ 模块 '{m}' 没有正向用例")
        if "负向" not in types and "异常" not in types:
            warnings.append(f"⚠️ 模块 '{m}' 没有负向/异常用例")

    # 检查异常场景关键词覆盖
    all_text = " ".join(
        (c.get("title", "") + c.get("steps", "") + c.get("expected", ""))
        for c in cases
    ).lower()
    exception_keywords = {
        "网络异常": ["网络", "断网", "弱网", "离线"],
        "超时": ["超时", "timeout"],
        "并发": ["并发", "重复", "幂等"],
        "权限": ["权限", "越权", "未登录", "token"],
        "数据边界": ["边界", "最大", "最小", "空", "特殊字符"],
    }
    for cat, kws in exception_keywords.items():
        if not any(kw.lower() in all_text for kw in kws):
            warnings.append(f"⚠️ 用例中未见 '{cat}' 场景相关关键词")

    return warnings


def main():
    if len(sys.argv) != 3:
        print("用法: python generate_test_cases.py <cases.json> <output.xlsx>")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 健康检查
    warnings = validate_cases(data)
    if warnings:
        print("=== 用例健康检查 ===")
        for w in warnings:
            print(w)
        print("===================")

    wb = Workbook()
    write_test_cases_sheet(wb, data)
    write_summary_sheet(wb, data)
    write_compatibility_sheet(wb, data)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    n = len(data.get("test_cases", []))
    print(f"✅ 已生成 {n} 条用例 → {output_path}")


if __name__ == "__main__":
    main()

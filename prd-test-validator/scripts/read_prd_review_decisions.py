#!/usr/bin/env python3
"""
read_prd_review_decisions.py
读取用户已经确认的 PRD 评审 Excel，输出结构化决定，用于驱动阶段 A 的用例生成。

使用方式:
    python read_prd_review_decisions.py <reviewed_xlsx>

输出 JSON 到 stdout，包含：
    - accepted: 用户接受默认建议的问题列表（含原始建议的新用例）
    - modified: 用户修改了建议的问题（含用户补充说明）
    - rejected: 用户拒绝处理的问题
    - pm_pending: 标记为"追问PM"的问题（无论 decision 如何，都需向产品反馈）
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook


def main():
    if len(sys.argv) != 2:
        print("用法: python read_prd_review_decisions.py <reviewed.xlsx>", file=sys.stderr)
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    wb = load_workbook(xlsx_path, data_only=True)
    
    # 找 PRD评审 sheet（先找精确匹配，避免被"评审摘要"抢先匹配）
    ws = None
    if "PRD评审" in wb.sheetnames:
        ws = wb["PRD评审"]
    else:
        # 兜底：不含"摘要"/"summary"的其他评审类 sheet
        for name in wb.sheetnames:
            if ("评审" in name or "review" in name.lower()) \
                and "摘要" not in name and "summary" not in name.lower():
                ws = wb[name]
                break
    if ws is None:
        print(json.dumps({"error": "找不到 PRD评审 sheet"}, ensure_ascii=False))
        sys.exit(1)

    # 表头在第 2 行
    headers = [cell.value for cell in ws[2]]
    header_map = {h: idx for idx, h in enumerate(headers) if h}

    required = ["用户决定", "问题 ID", "严重等级", "维度", "PRD 位置",
                "问题说明", "默认处理建议", "建议详情", "若确认将新增的用例"]
    missing = [c for c in required if c not in header_map]
    if missing:
        print(json.dumps({"error": f"表头缺少: {missing}"}, ensure_ascii=False))
        sys.exit(1)

    accepted = []
    modified = []
    rejected = []
    pm_pending = []
    undecided = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or all(c is None for c in row):
            continue

        decision = str(row[header_map["用户决定"]] or "").strip()
        issue_id = str(row[header_map["问题 ID"]] or "").strip()
        if not issue_id:
            continue

        severity_display = str(row[header_map["严重等级"]] or "").strip()
        dimension = str(row[header_map["维度"]] or "").strip()
        location = str(row[header_map["PRD 位置"]] or "").strip()
        problem = str(row[header_map["问题说明"]] or "").strip()
        default_action = str(row[header_map["默认处理建议"]] or "").strip()
        default_detail = str(row[header_map["建议详情"]] or "").strip()
        new_cases_raw = str(row[header_map["若确认将新增的用例"]] or "").strip()
        user_notes = ""
        if "用户补充/覆盖说明" in header_map:
            user_notes = str(row[header_map["用户补充/覆盖说明"]] or "").strip()

        # 解析新用例
        new_cases = []
        if new_cases_raw and new_cases_raw != "（无新增用例）":
            for line in new_cases_raw.splitlines():
                line = line.strip().lstrip("•").strip()
                if line:
                    new_cases.append(line)

        item = {
            "id": issue_id,
            "severity": severity_display,
            "dimension": dimension,
            "location": location,
            "problem": problem,
            "default_action": default_action,
            "default_detail": default_detail,
            "new_cases": new_cases,
            "user_notes": user_notes,
        }

        if decision == "接受":
            accepted.append(item)
        elif decision == "修改":
            modified.append(item)
        elif decision == "拒绝":
            rejected.append(item)
        else:
            undecided.append(item)

        # "追问PM"类型，不论决定如何都标记一份副本
        if "追问" in default_action or "PM" in default_action:
            pm_pending.append(item)

    result = {
        "accepted_count": len(accepted),
        "modified_count": len(modified),
        "rejected_count": len(rejected),
        "pm_pending_count": len(pm_pending),
        "undecided_count": len(undecided),
        "accepted": accepted,
        "modified": modified,
        "rejected": rejected,
        "pm_pending": pm_pending,
        "undecided": undecided,
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

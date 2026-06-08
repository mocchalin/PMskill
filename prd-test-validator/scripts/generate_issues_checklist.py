#!/usr/bin/env python3
"""
generate_issues_checklist.py
把审查报告中识别出的问题，输出为带勾选列的 Excel 清单。
用户在勾选列填 √（或 x、yes、1）后另存，再作为 generate_fix_prompts.py 的输入。

使用方式:
    python generate_issues_checklist.py <issues_json> <output_xlsx>

输入 JSON schema:
{
    "project": "项目名",
    "code_reviewed": ["src/app.py", "src/auth.py"],
    "issues": [
        {
            "id": "ISSUE-001",
            "severity": "critical",        // critical / major / minor
            "dimension": "安全",            // 需求覆盖/代码质量/安全/性能并发
            "file": "src/app.py",
            "line": "34-35",
            "title": "SQL 注入：INSERT 语句字符串拼接",
            "code_snippet": "cursor.execute(f\"INSERT INTO files ... VALUES ('{filename}', ...)\")",
            "description": "所有 SQL 都用 f-string 拼接用户输入...",
            "fix_suggestion": "改用参数化查询：cursor.execute(\"INSERT INTO files ... VALUES (?, ...)\", (filename, ...))",
            "related_ac": "AC-2.1.1",
            "related_tc": "TC-UP-006"
        }
    ]
}
"""

import json
import sys
from pathlib import Path
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="305496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

SEVERITY_FILLS = {
    "critical": PatternFill("solid", start_color="FFC7CE"),
    "major": PatternFill("solid", start_color="FFEB9C"),
    "minor": PatternFill("solid", start_color="C6EFCE"),
}

SEVERITY_DISPLAY = {
    "critical": "🔴 严重",
    "major": "🟡 一般",
    "minor": "🟢 建议",
}

COLUMNS = [
    ("修复", "fix_flag", 8),           # 用户勾选列
    ("问题 ID", "id", 12),
    ("严重等级", "severity_display", 12),
    ("维度", "dimension", 12),
    ("文件", "file", 24),
    ("行号", "line", 10),
    ("问题标题", "title", 32),
    ("代码片段", "code_snippet", 40),
    ("问题说明", "description", 40),
    ("修复建议", "fix_suggestion", 40),
    ("关联 AC", "related_ac", 14),
    ("关联用例", "related_tc", 14),
]


def write_issues_sheet(wb, data):
    ws = wb.active
    ws.title = "问题清单"

    # 顶部说明行
    ws["A1"] = (
        "✍️ 使用说明：在【修复】列填入 √ (或 x / yes / 1) 勾选想修复的问题，另存 Excel 后运行 generate_fix_prompts.py"
    )
    ws["A1"].font = Font(name="Arial", size=10, italic=True, color="C00000")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    # 表头（行 2）
    for col_idx, (header, _, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    issues = data.get("issues", [])

    for row_offset, issue in enumerate(issues):
        row_idx = row_offset + 3
        severity = issue.get("severity", "").lower()

        values = {
            "fix_flag": "",
            "id": issue.get("id", ""),
            "severity_display": SEVERITY_DISPLAY.get(severity, severity),
            "dimension": issue.get("dimension", ""),
            "file": issue.get("file", ""),
            "line": issue.get("line", ""),
            "title": issue.get("title", ""),
            "code_snippet": issue.get("code_snippet", ""),
            "description": issue.get("description", ""),
            "fix_suggestion": issue.get("fix_suggestion", ""),
            "related_ac": issue.get("related_ac", ""),
            "related_tc": issue.get("related_tc", ""),
        }

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=values[key])
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER

            if key == "severity_display":
                fill = SEVERITY_FILLS.get(severity)
                if fill:
                    cell.fill = fill
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    cell.font = Font(name="Arial", size=10, bold=True)

            if key in ("id", "dimension", "line", "related_ac", "related_tc"):
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            if key == "fix_flag":
                # 淡黄高亮，提示用户这是交互列
                cell.fill = PatternFill("solid", start_color="FFF2CC")
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

    # 列宽
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 行高
    for row_idx in range(3, len(issues) + 3):
        ws.row_dimensions[row_idx].height = 80

    # 冻结表头（首列修复也随列滚，但冻住到 B 列方便看 ID）
    ws.freeze_panes = "C3"

    # 筛选器
    last_col_letter = get_column_letter(len(COLUMNS))
    last_row = len(issues) + 2
    ws.auto_filter.ref = f"A2:{last_col_letter}{last_row}"

    # 数据校验：修复列提供下拉选项 √ 或空
    if issues:
        dv = DataValidation(
            type="list",
            formula1='"√,x"',
            allow_blank=True,
            showDropDown=False,  # False = 显示下拉箭头
        )
        dv.prompt = "填 √ 表示需要修复"
        dv.promptTitle = "是否修复"
        dv.add(f"A3:A{len(issues) + 2}")
        ws.add_data_validation(dv)


def write_summary_sheet(wb, data):
    ws = wb.create_sheet("统计", 0)
    issues = data.get("issues", [])

    ws["A1"] = f"代码审查问题清单 - {data.get('project', '未命名')}"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="305496")
    ws.merge_cells("A1:D1")

    reviewed = data.get("code_reviewed", [])
    ws["A2"] = f"审查文件：{', '.join(reviewed) if reviewed else '未指定'}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="595959")
    ws.merge_cells("A2:D2")

    total = len(issues)
    by_severity = Counter(i.get("severity", "").lower() for i in issues)
    by_dimension = Counter(i.get("dimension", "") for i in issues)
    by_file = Counter(i.get("file", "") for i in issues)

    row = 4
    ws.cell(row=row, column=1, value="总览").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    ws.cell(row=row, column=1, value="问题总数").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total)
    row += 2

    def write_block(title, counter, color_map=None):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        ws.cell(row=row, column=1, value="分类").font = Font(bold=True)
        ws.cell(row=row, column=2, value="数量").font = Font(bold=True)
        ws.cell(row=row, column=3, value="占比").font = Font(bold=True)
        row += 1
        for k, v in counter.most_common():
            if not k:
                continue
            display = SEVERITY_DISPLAY.get(k, k) if title == "按严重等级" else k
            ws.cell(row=row, column=1, value=display)
            ws.cell(row=row, column=2, value=v).alignment = Alignment(horizontal="center")
            pct = f"{(v / total * 100):.1f}%" if total else "0%"
            ws.cell(row=row, column=3, value=pct).alignment = Alignment(horizontal="center")
            if color_map and k in color_map:
                ws.cell(row=row, column=1).fill = color_map[k]
            row += 1
        row += 1

    write_block("按严重等级", by_severity, SEVERITY_FILLS)
    write_block("按维度", by_dimension)
    write_block("按文件", by_file)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20


def main():
    if len(sys.argv) != 3:
        print("用法: python generate_issues_checklist.py <issues.json> <output.xlsx>")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    write_issues_sheet(wb, data)
    write_summary_sheet(wb, data)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    n = len(data.get("issues", []))
    print(f"✅ 已生成 {n} 条问题清单 → {output_path}")
    print("   用户下一步：在 Excel 的【修复】列填 √ 勾选要修的问题，另存后作为 generate_fix_prompts.py 的输入")


if __name__ == "__main__":
    main()
